# -*- coding: utf-8 -*-
"""Extract the OTK checklist template from the source sheet into JSON,
replacing formula-driven cells with substitution tokens."""
import json, os
import openpyxl

SRC = os.path.join(os.path.dirname(__file__), '..', '..', 'KTPN_RUVN_Busbar_Updated.xlsx')
DATA = os.path.join(os.path.dirname(__file__), '..', 'src', 'KtpnConfigurator.App', 'Data')
wb = openpyxl.load_workbook(SRC, data_only=False)
ws = wb['Чек-лист ОТК']

# Rows with formulas get mapped to tokens replaced at render time.
DYNAMIC = {
    15: 'Габариты ШхДхВ: {width}x{length}x{height} мм',
    16: 'Металл корпуса: {steelType} {thickness} мм',
    25: 'Цвет корпуса: {bodyColor}',
    37: 'Цвет дверей: {doorColor}',
}
# Section header rows (column B empty)
sections = []
current = None
for r in range(7, ws.max_row + 1):
    a = ws.cell(r, 1).value
    b = ws.cell(r, 2).value
    if a is None:
        continue
    if r in DYNAMIC:
        text = DYNAMIC[r]
    else:
        text = str(a)
    if b is None:  # header
        current = {'name': text, 'items': []}
        sections.append(current)
    else:
        if current is None:
            current = {'name': '', 'items': []}
            sections.append(current)
        current['items'].append(text)

doc = {
    'qcChecklist': {
        'title': 'Акт осмотра КТПН №_______',
        'committee': 'Состав комиссии: Инженер по качеству Ершов А.М. / _________________________',
        'sections': sections,
    },
    'productionOrder': {
        'header': 'В работу\nКонструкторская служба\nОтдел Снабжения',
        'orderLine': 'Приказ № _______ от ___.___.202_ г.',
        'subtitle': 'На запуск в производство электрооборудования',
        'signatures': 'Подписи:  ________________ / М.М. Жамалетдинов        ________________ / М.Д. Богатенков',
    },
}
with open(os.path.join(DATA, 'doc_templates.json'), 'w', encoding='utf-8') as f:
    json.dump(doc, f, ensure_ascii=False, indent=2)
print('doc_templates.json written:', len(sections), 'sections,',
      sum(len(s['items']) for s in sections), 'items')
for s in sections:
    print(' ', s['name'], '->', len(s['items']), 'items')
