# -*- coding: utf-8 -*-
"""Extracts reference data from the source workbook into clean JSON catalogs
and generates golden test vectors by re-implementing the Excel formulas exactly.
This is the independent oracle the C# CalculationEngine is tested against."""
import json, math, os, sys
import openpyxl

SRC = os.path.join(os.path.dirname(__file__), 'source', 'KTPN_RUVN_Busbar_Updated.xlsx')
DATA = os.path.join(os.path.dirname(__file__), '..', 'src', 'KtpnConfigurator.App', 'Data')
os.makedirs(DATA, exist_ok=True)

wb = openpyxl.load_workbook(SRC, data_only=False)

def col(ws, c, r1, r2):
    return [ws.cell(r, c).value for r in range(r1, r2 + 1)]

# ---------- Transformers (consolidated DB is the single source of truth) ----------
ws = wb['БД_Трансформаторы']
transformers = []
for r in range(2, ws.max_row + 1):
    mark = ws.cell(r, 1).value
    if not mark:
        continue
    transformers.append({
        'mark': mark,
        'manufacturer': ws.cell(r, 2).value,
        'powerKva': ws.cell(r, 3).value,
        'lengthMm': ws.cell(r, 4).value,
        'widthMm': ws.cell(r, 5).value,
        'heightMm': ws.cell(r, 6).value,
        'massKg': ws.cell(r, 7).value,
        'ratedCurrentA': ws.cell(r, 8).value,
        'busbarHv': ws.cell(r, 9).value,
        'busbarLv': ws.cell(r, 10).value,
        'busbarPe': ws.cell(r, 11).value,
    })

# ---------- Apparatus DB ----------
ws = wb['БД_Аппараты_РУНН']
apparatus = []
for r in range(2, ws.max_row + 1):
    if not ws.cell(r, 1).value:
        continue
    apparatus.append({
        'type': ws.cell(r, 1).value,
        'manufacturer': ws.cell(r, 2).value,
        'country': ws.cell(r, 3).value,
        'series': ws.cell(r, 4).value,
        'currentRange': ws.cell(r, 5).value,
    })

# ---------- Lists sheet ----------
L = wb['Списки']

def named(c, r1, r2):
    return [v for v in col(L, c, r1, r2) if v is not None]

# channels: AA2:AB13 (size -> weight per m)
channels = []
for r in range(2, 14):
    channels.append({'size': L.cell(r, 27).value, 'weightPerM': L.cell(r, 28).value})
# steels: Y2:Z5 (thickness -> weight per m2)
steels = []
for r in range(2, 6):
    steels.append({'thicknessMm': float(L.cell(r, 25).value), 'weightPerM2': L.cell(r, 26).value})

options = {
    'steelTypes': named(9, 2, 3),                 # I
    'steelThicknesses': [float(x) for x in named(10, 2, 5)],  # J
    'channels': channels,
    'steels': steels,
    'ralColors': named(12, 2, 10),                # L
    'gridCompanies': named(13, 2, 6),             # M
    'voltages': named(22, 2, 3),                  # V
    'ruvnTypes': named(19, 2, 4),                 # S
    'ruvnSwitches': named(20, 2, 4),              # T
    'ruvnNominals': [int(float(x)) for x in named(21, 2, 4)],   # U 400/630/1000
    'fuseTypes': named(23, 2, 5),                 # W
    'fuseNominals': named(24, 2, 16),             # X
    'cableExecutions': ['Кабельный', 'Воздушный'], # Hardcoded as it's not in the lists
    # Cleaned current rating ряд for ПВР/РЕ/АВ (decision per §11 of the spec):
    'lvNominals': [400, 630, 1000, 1600, 2000, 2500, 3150, 4000],
    # Cleaned TT ratios:
    'ttRatios': ['300/5', '400/5', '600/5', '800/5', '1000/5', '1500/5', '2000/5', '2500/5'],
    'pvrManufacturers': named(14, 2, 6),          # N
    'reManufacturers': named(15, 2, 5),           # O
    'rpsManufacturers': named(16, 2, 4),          # P
    'avManufacturers': named(17, 2, 7),           # Q
    'yesNo': named(18, 2, 3),                     # R
    'methodology': {
        'floorSheetKgPerM2': 35.0,
        'frameCoef': 1.15,
        'bodyWasteCoef': 1.25,
        'standardLengths': [2000, 3000, 3500],
        'minDimensionMm': 1800,
        'heightThresholdMm': 1800,
        'tallHeightMm': 2400,
        'shortHeightMm': 2250,
        'passageWidthMm': 2420,
        'multiBayWidthMm': 2420,
        'standardWidthMm': 2000,
        'multiBayAvQty': 8,
        'multiBayRpsQty': 4,
    },
}

with open(os.path.join(DATA, 'transformers.json'), 'w', encoding='utf-8') as f:
    json.dump(transformers, f, ensure_ascii=False, indent=2)
with open(os.path.join(DATA, 'apparatus.json'), 'w', encoding='utf-8') as f:
    json.dump(apparatus, f, ensure_ascii=False, indent=2)
with open(os.path.join(DATA, 'options.json'), 'w', encoding='utf-8') as f:
    json.dump(options, f, ensure_ascii=False, indent=2)

print('Catalogs written:', len(transformers), 'transformers,', len(apparatus), 'apparatus')

# ================= ORACLE: re-implementation of Excel formulas =================
M = options['methodology']
chan_w = {c['size']: c['weightPerM'] for c in channels}
steel_w = {s['thicknessMm']: s['weightPerM2'] for s in steels}
tmap = {t['mark']: t for t in transformers}

def xround(x, n=0):
    f = 10 ** n
    return math.floor(abs(x) * f + 0.5) / f * (1 if x >= 0 else -1)

def calc_length(ruvn_type, len_ruvn, tr_len, len_runn, buffer):
    base = (0 if ruvn_type == 'Нет' else len_ruvn) + tr_len + len_runn + buffer
    if base <= M['standardLengths'][0]: return M['standardLengths'][0]
    if base <= M['standardLengths'][1]: return M['standardLengths'][1]
    if base <= M['standardLengths'][2]: return M['standardLengths'][2]
    return base

def with_min_override(calc, manual):
    if manual is not None:
        if manual == 0: return 0
        return max(manual, M['minDimensionMm'])
    return calc

def calc_width(ruvn_type, av_on, av_qty, rps_on, rps_qty, tr_w, tol):
    if ruvn_type == 'Проходная': return M['passageWidthMm']
    multi = ((av_qty if av_on else 0) > M['multiBayAvQty']) or ((rps_qty if rps_on else 0) > M['multiBayRpsQty'])
    base_min = M['multiBayWidthMm'] if multi else M['standardWidthMm']
    return max(base_min, tr_w + tol)

def calc_height(tr_h):
    return M['tallHeightMm'] if tr_h >= M['heightThresholdMm'] else M['shortHeightMm']

def base_mass(L_, W_, channel):
    cw = chan_w[channel]
    return xround(((L_ * 2 + W_ * 4) / 1000) * cw * M['frameCoef'] + (L_ * W_ / 1_000_000) * M['floorSheetKgPerM2'], 0)

def body_mass(L_, W_, H_, thickness):
    sw = steel_w[thickness]
    area = ((L_ + W_) * 2 * H_ / 1_000_000) + (L_ * W_ / 1_000_000)
    return xround(area * sw * M['bodyWasteCoef'], 0)

def evaluate(cfg):
    t = tmap[cfg['mark']]
    L_calc = calc_length(cfg['ruvnType'], cfg['lenRuvn'], t['lengthMm'], cfg['lenRunn'], cfg['buffer'])
    L_fin = with_min_override(L_calc, cfg.get('manualLength'))
    W_calc = calc_width(cfg['ruvnType'], cfg['avOn'], cfg['avQty'], cfg['rpsOn'], cfg['rpsQty'], t['widthMm'], cfg['tol'])
    W_fin = with_min_override(W_calc, cfg.get('manualWidth'))
    H_calc = calc_height(t['heightMm'])
    H_fin = cfg.get('manualHeight') if cfg.get('manualHeight') is not None else H_calc
    bm = base_mass(L_fin, W_fin, cfg['channel'])
    cm = body_mass(L_fin, W_fin, H_fin, cfg['thickness'])
    gross = t['massKg'] + bm + cm
    nom_in = max(cfg['pvrNom'] if cfg['pvrOn'] else 0,
                 cfg['reNom'] if cfg['reOn'] else 0,
                 cfg['avInNom'] if cfg['avInOn'] else 0)
    ok = nom_in >= t['ratedCurrentA']
    return {
        'mark': cfg['mark'], 'ratedCurrentA': t['ratedCurrentA'],
        'lengthCalc': L_calc, 'lengthFinal': L_fin,
        'widthCalc': W_calc, 'widthFinal': W_fin,
        'heightCalc': H_calc, 'heightFinal': H_fin,
        'baseMass': bm, 'bodyMass': cm, 'grossMass': gross,
        'inputNominal': nom_in, 'validationOk': ok,
    }

def base_cfg(**over):
    c = dict(mark='ТМГ-400 (Алагеум)', ruvnType='Тупиковая', lenRuvn=1300, lenRunn=600,
             tol=300, buffer=10, channel='10П', thickness=2.0,
             avOn=True, avQty=2, rpsOn=False, rpsQty=0,
             pvrOn=True, pvrNom=630, reOn=False, reNom=630, avInOn=False, avInNom=630,
             manualLength=None, manualWidth=None, manualHeight=None)
    c.update(over); return c

cases = {
    'default_TMG400_Alageum': base_cfg(),
    'small_TMG25_short_height': base_cfg(mark='ТМГ-25', avQty=2),
    'passage_width_2420': base_cfg(ruvnType='Проходная'),
    'no_ruvn_compartment': base_cfg(ruvnType='Нет'),
    'many_av_multibay': base_cfg(avQty=10),
    'many_rps_multibay': base_cfg(rpsOn=True, rpsQty=5),
    'large_TMG2500_tall': base_cfg(mark='ТМГ-2500 (Алагеум)', pvrNom=4000),
    'manual_overrides': base_cfg(manualLength=4200, manualWidth=2500, manualHeight=2600),
    'manual_zero_length': base_cfg(manualLength=0),
    'manual_below_min': base_cfg(manualLength=1500, manualWidth=1000),
    'validation_fail': base_cfg(mark='ТМГ-1600 (Алагеум)', pvrNom=630),
    'validation_pass_via_re': base_cfg(pvrOn=False, reOn=True, reNom=2500, mark='ТМГ-1000 (Алагеум)'),
}
golden = {name: evaluate(cfg) for name, cfg in cases.items()}
golden_cfgs = {name: cfg for name, cfg in cases.items()}

out = {'configs': golden_cfgs, 'expected': golden}
with open(os.path.join(os.path.dirname(__file__), 'golden.json'), 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

print('\n=== GOLDEN VECTORS ===')
for name in cases:
    print(name, '->', json.dumps(golden[name], ensure_ascii=False))
